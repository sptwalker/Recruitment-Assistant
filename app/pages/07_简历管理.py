"""简历分析管理模块 — 3 Tab 页面。

Tab 1: 简历自动解析入库（扫描 → 提取文本 → AI 结构化 → 去重 → 入库）
Tab 2: 简历库浏览（搜索 / 详情 / 删除 / 屏蔽 / 面试邀约）
Tab 3: 招聘岗位录入/匹配（录入岗位 → AI 匹配候选人）
"""

import importlib
import io
import os
import re
import time as _time
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import streamlit as st
from loguru import logger

from components.layout import inject_vibe_style, page_header
from recruitment_assistant.config.settings import get_settings
import recruitment_assistant.parsers.pdf_resume_parser as resume_parser_module
from recruitment_assistant.schemas.resume_archive import ResumeSourceCreate
import recruitment_assistant.services.resume_ai_service as resume_ai_service_module
from recruitment_assistant.services.resume_archive_service import ResumeArchiveService
from recruitment_assistant.services.job_service import JobService
from recruitment_assistant.storage.db import create_session
from recruitment_assistant.storage.resume_db import create_resume_session, init_resume_database
from recruitment_assistant.storage.models import JobPosition

resume_parser_module = importlib.reload(resume_parser_module)
extract_doc_text = resume_parser_module.extract_doc_text
extract_text_from_docx = resume_parser_module.extract_text_from_docx
extract_text_from_pdf = resume_parser_module.extract_text_from_pdf
is_empty_or_corrupted = resume_parser_module.is_empty_or_corrupted
has_garbled_text = resume_parser_module.has_garbled_text

resume_ai_service_module = importlib.reload(resume_ai_service_module)
ResumeAIService = resume_ai_service_module.ResumeAIService
normalize_platform = resume_ai_service_module.normalize_platform

init_resume_database()
get_settings.cache_clear()
settings = get_settings()

st.set_page_config(page_title="简历管理", layout="wide", initial_sidebar_state="collapsed")
inject_vibe_style("简历管理")
page_header("简历分析管理", "自动解析入库、浏览管理、岗位匹配。面试跟进请到「面试管理」。")


@st.cache_resource
def get_ai_service(api_key: str, base_url: str, model: str) -> ResumeAIService:
    return ResumeAIService(
        api_key=api_key,
        base_url=base_url,
        model=model,
    )


ai_service = get_ai_service(settings.ai_api_key, settings.ai_base_url, settings.ai_model)

RESUME_DIRS = {
    "BOSS直聘": settings.attachment_dir / "boss",
    "智联招聘": settings.attachment_dir / "zhilian",
    "51前程无忧": settings.attachment_dir / "51job",
}


def scan_resume_files() -> list[dict]:
    """扫描三个平台目录下所有 PDF/DOCX 简历文件。

    带「（附件作品）」标记的文件不是独立简历，是某候选人的作品附件，跳过；
    解析入库主流程结束后，再由 `link_attachment_works_for_resume` 按文件名配对到对应候选人。
    """
    files = []
    for platform, base_dir in RESUME_DIRS.items():
        if not base_dir.exists():
            continue
        for f in sorted(base_dir.rglob("*"), key=lambda p: p.stat().st_mtime, reverse=True):
            if f.suffix.lower() in (".pdf", ".docx", ".doc") and f.is_file():
                if "（附件作品）" in f.stem:
                    continue
                files.append({
                    "path": f,
                    "platform": platform,
                    "name": f.name,
                    "size_kb": round(f.stat().st_size / 1024, 1),
                    "mtime": datetime.fromtimestamp(f.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
                })
    return files


def find_attachment_works_path_for(resume_path: Path) -> str | None:
    """根据简历文件名在同目录找配对的作品 PDF。

    简历命名：<name>-<age>-<edu>[-position]-<source>-<date>-<time>-<seq>.pdf
    作品命名：<name>-<age>-<edu>[-position]-<source>-（附件作品）-<date>-<time>-<seq>.pdf

    规则：同目录、文件名以「<resume_stem 前缀（去末尾日期/时间/序号 3 段）>」为前缀、含「（附件作品）」的 .pdf 文件。
    """
    if not resume_path.exists():
        return None
    folder = resume_path.parent
    stem = resume_path.stem
    parts = stem.split("-")
    # 至少 3 段（姓名/年龄/学历/...日期/时间/序号），剥掉末尾 3 段当 prefix。
    if len(parts) < 4:
        return None
    prefix = "-".join(parts[:-3])
    for f in folder.glob("*.pdf"):
        if f == resume_path:
            continue
        if "（附件作品）" not in f.stem:
            continue
        if f.stem.startswith(prefix + "-"):
            return str(f)
    return None


def extract_text(path: Path) -> str:
    """根据文件类型提取纯文本。"""
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return extract_text_from_pdf(path)
    elif suffix == ".docx":
        return extract_text_from_docx(path)
    elif suffix == ".doc":
        return extract_doc_text(path)
    return ""


def should_skip_empty_or_corrupted(path: Path) -> bool:
    """页面侧最终判定：老 .doc 优先用当前进程内重载后的提取器检测，避免旧缓存误判。"""
    if path.suffix.lower() == ".doc":
        try:
            return len(extract_doc_text(path).strip()) < 50
        except Exception:
            return True
    return is_empty_or_corrupted(path)


def infer_candidate_from_filename(file_name: str) -> tuple[str | None, int | None, str | None]:
    """从标准附件名中提取姓名/年龄/学历，用于 AI 调用前预去重。"""
    stem = Path(file_name).stem
    parts = [part.strip() for part in re.split(r"[-_｜|]", stem) if part.strip()]
    name = parts[0] if parts else None
    age = None
    education_level = None
    for part in parts[1:5]:
        age_match = re.search(r"(\d{1,2})\s*岁", part)
        if age_match:
            age = int(age_match.group(1))
        if part in {"高中", "中专", "大专", "专科", "本科", "硕士", "研究生", "博士"}:
            education_level = "硕士" if part == "研究生" else ("大专" if part == "专科" else part)
    if name and not re.search(r"[\u4e00-\u9fffA-Za-z]", name):
        name = None
    return name, age, education_level


def filter_duplicate_files_before_ai(files: list[dict]) -> tuple[list[dict], list[str]]:
    """在调用 AI 前用文件名里的姓名/年龄/学历先查库去重，减少 AI 流量消耗。"""
    if not files:
        return [], []
    kept: list[dict] = []
    skipped: list[str] = []
    session = create_resume_session()
    try:
        svc = ResumeArchiveService(session)
        for file_info in files:
            name, age, education_level = infer_candidate_from_filename(file_info["name"])
            if name and svc.is_duplicate(name=name, age=age, education_level=education_level):
                skipped.append(f"{name}（{file_info['name']}）")
            else:
                kept.append(file_info)
    finally:
        session.close()
    return kept, skipped


# ==================== 3 Tab 页面 ====================
tabs = st.tabs([
    "📥 简历自动解析入库",
    "📋 简历库浏览",
    "🎯 招聘岗位录入/匹配",
])

# ==================== Tab 1: 简历自动解析入库 ====================
with tabs[0]:
    st.markdown("### 简历自动解析入库")

    if not ai_service.is_configured:
        st.warning("⚠️ AI API Key 未配置。请到「平台登录 → AI模型」标签配置 `AI_API_KEY` / `AI_BASE_URL` / `AI_MODEL`，保存后重启 Streamlit。")

    # 数据库当前状态（非任务信息，保留在窗口外）
    all_files = scan_resume_files()
    session = create_resume_session()
    svc = ResumeArchiveService(session)
    stats = svc.get_stats()
    session.close()

    total = stats["total"] or 0
    boss_count = stats["platform_counts"].get("BOSS直聘", 0)
    zhilian_count = stats["platform_counts"].get("智联招聘", 0)
    qiancheng_count = stats["platform_counts"].get("51前程无忧", 0)

    def _ring_svg(value: int, total_n: int, color: str) -> str:
        pct = round(value / total_n * 100) if total_n > 0 else 0
        r, circ = 12, 75.4
        offset = circ * (1 - pct / 100)
        return (
            f'<svg width="30" height="30" viewBox="0 0 30 30">'
            f'<circle cx="15" cy="15" r="{r}" fill="none" stroke="#e5e7eb" stroke-width="2.5"/>'
            f'<circle cx="15" cy="15" r="{r}" fill="none" stroke="{color}" stroke-width="2.5" '
            f'stroke-dasharray="{circ}" stroke-dashoffset="{offset:.1f}" '
            f'transform="rotate(-90 15 15)" stroke-linecap="round"/>'
            f'<text x="15" y="15" text-anchor="middle" dominant-baseline="central" '
            f'font-size="8" font-weight="700" fill="{color}">{pct}%</text>'
            f'</svg>'
        )

    def _stat_card(label: str, value: int, ring_html: str = "", is_primary: bool = False) -> str:
        ring_part = f'<div style="flex-shrink:0;margin-left:8px;">{ring_html}</div>' if ring_html else ''
        label_style = "font-size:16px;font-weight:900;font-family:SimHei,sans-serif;" if is_primary else "font-size:15px;font-weight:700;"
        return (
            f'<div style="background:var(--color-surface);border:1px solid var(--color-border);border-radius:12px;'
            f'padding:10px 14px;display:flex;align-items:center;justify-content:center;min-height:54px;">'
            f'<div style="display:flex;align-items:baseline;gap:6px;">'
            f'<span style="{label_style}color:var(--color-text-secondary);white-space:nowrap;">{label}：</span>'
            f'<span style="font-size:24px;font-weight:800;color:var(--color-text);line-height:1;">{value}</span>'
            f'</div>{ring_part}</div>'
        )

    col0, col1, col2, col3, col4 = st.columns(5)
    col0.markdown(_stat_card("简历库总数", total, is_primary=True), unsafe_allow_html=True)
    col1.markdown(_stat_card("待入库文件", len(all_files)), unsafe_allow_html=True)
    col2.markdown(_stat_card("BOSS直聘", boss_count, _ring_svg(boss_count, total, "#0a7d2e")), unsafe_allow_html=True)
    col3.markdown(_stat_card("智联招聘", zhilian_count, _ring_svg(zhilian_count, total, "#1d4ed8")), unsafe_allow_html=True)
    col4.markdown(_stat_card("51前程无忧", qiancheng_count, _ring_svg(qiancheng_count, total, "#b45309")), unsafe_allow_html=True)

    st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)

    # ---------- 任务状态：用 session_state 跨 rerun 维护 ----------
    # parse_task_state: "idle" / "running" / "stopping"
    # parse_queue:      待处理文件列表（任务开始时快照，运行中不变）
    # parse_index:      当前处理到第几个（每份 rerun 一次后 +1）
    # parse_results:    汇总统计与明细
    if "parse_task_state" not in st.session_state:
        st.session_state.parse_task_state = "idle"
    if "parse_log_lines" not in st.session_state:
        st.session_state.parse_log_lines = []
    if "parse_queue" not in st.session_state:
        st.session_state.parse_queue = []
    if "parse_index" not in st.session_state:
        st.session_state.parse_index = 0
    if "parse_since_date" not in st.session_state:
        st.session_state.parse_since_date = date.today() - timedelta(days=7)
    if "parse_results" not in st.session_state:
        st.session_state.parse_results = {
            "success_count": 0,
            "skip_count": 0,
            "fail_count": 0,
            "success_files": [],
            "skip_files": [],
            "failed_files": {
                "空白/损坏": [],
                "文本过短": [],
                "AI 解析异常": [],
                "AI 返回空结果": [],
                "入库失败": [],
            },
        }

    # all_files 已在上方 metric 区扫描，运行中用 parse_queue 快照

    def append_scan_summary(lines: list[str], files: list[dict]) -> None:
        ts = datetime.now().strftime("%H:%M:%S")
        by_platform: dict[str, int] = {}
        for f in files:
            by_platform[f["platform"]] = by_platform.get(f["platform"], 0) + 1
        lines.append(f"[{ts}] 📂 扫描简历目录")
        lines.append(f"  共扫描到 {len(files)} 个简历文件")
        for p in ("BOSS直聘", "智联招聘", "51前程无忧"):
            lines.append(f"    - {p}: {by_platform.get(p, 0)} 份")
        if not files:
            lines.append("  ℹ️ 三个平台目录下暂无简历文件")
        else:
            since_repr = st.session_state.get(
                "parse_since_date", date.today() - timedelta(days=7)
            ).isoformat()
            lines.append(f"  ⏰ 当前日期过滤：≥ {since_repr}（可在按钮行调整）")
            if ai_service.is_configured:
                lines.append("  👉 点击下方「自动解析入库」按钮启动任务")
            else:
                lines.append("  ⚠️ AI 未配置，「开始」按钮已禁用")
        lines.append("")

    # 第一次进入时自动追加扫描统计
    if not st.session_state.parse_log_lines:
        append_scan_summary(st.session_state.parse_log_lines, all_files)

    # ---------- 渲染日志窗口（HTML，文字可选中，无灰化）----------
    import html as _html
    import streamlit.components.v1 as components

    def render_log_window() -> None:
        # 只保留最后 500 行展示，避免 DOM 过大
        recent = st.session_state.parse_log_lines[-500:]
        # 按追加顺序渲染：旧信息在上，新信息继续向下输出
        def _log_row_class(line: str) -> str:
            if any(token in line for token in ("❌", "⚠️", "失败", "跳过", "异常", "损坏", "过短")):
                return "log-row log-row-error"
            return "log-row"

        escaped_rows = "".join(
            f"<div class='{_log_row_class(line)}'>{_html.escape(line) or '&nbsp;'}</div>"
            for line in recent
        )
        # 用 components.html 而不是 st.markdown：markdown 会过滤 <script>，
        # 自动滚到底脚本必须靠真正的 iframe 执行。
        # overflow-y: scroll（不是 auto）= 始终保留滚动条占位，避免内容达阈值时整体宽度抖动。
        html = f"""
        <div id='resume-log-window' class='resume-log-window'>
          {escaped_rows}
        </div>
        <script>
          (function() {{
            const win = document.getElementById('resume-log-window');
            if (!win) return;
            // 首次渲染：直接到底
            win.scrollTop = win.scrollHeight;
            // 监听后续 DOM 变化（Streamlit 增量重渲染时新行追加进来即贴底）
            const obs = new MutationObserver(() => {{
              win.scrollTop = win.scrollHeight;
            }});
            obs.observe(win, {{ childList: true, subtree: true, characterData: true }});
          }})();
        </script>
        <style>
          html, body {{ margin: 0; padding: 0; background: transparent; }}
          .resume-log-window {{
            height: 460px;
            overflow-y: scroll;          /* 始终显示滚动条 */
            background: #ffffff;
            color: #262730;
            font-family: 'Consolas','Monaco','Microsoft YaHei Mono','Microsoft YaHei',monospace;
            font-size: 13px;
            line-height: 1.55;
            padding: 12px 14px;
            border-radius: 6px;
            border: 1px solid #e6eaf1;
            white-space: pre-wrap;
            word-break: break-all;
            user-select: text;            /* 允许选中文字 */
            scrollbar-gutter: stable;
          }}
          .resume-log-window .log-row {{
            color: #262730;
          }}
          .resume-log-window .log-row-error {{
            color: #c0392b;
            font-weight: 700;
          }}
          .resume-log-window::-webkit-scrollbar {{ width: 10px; }}
          .resume-log-window::-webkit-scrollbar-track {{ background: #f0f2f6; }}
          .resume-log-window::-webkit-scrollbar-thumb {{
            background: #c8ccd4;
            border-radius: 5px;
          }}
          .resume-log-window::-webkit-scrollbar-thumb:hover {{ background: #9aa0a6; }}
        </style>
        """
        # iframe 高度比 480 多 6px，给边框/圆角留余量
        components.html(html, height=486, scrolling=False)

    render_log_window()

    # 日志窗口与按钮区之间的视觉气口
    st.markdown("<div style='height: 18px'></div>", unsafe_allow_html=True)

    # ---------- 操作按钮区（窗口外不显示任何任务信息）----------
    task_state = st.session_state.parse_task_state
    is_idle = task_state == "idle"
    is_running = task_state == "running"
    is_stopping = task_state == "stopping"

    btn_cols = st.columns([1.8, 0.3, 1.2, 1.2, 1.2, 2.1])
    # 索引 0=日期 1=间隔 2=开始/停止 3=重新扫描 4=清空 5=进度

    with btn_cols[0]:
        st.markdown(
            "<div style='font-size:14px; font-weight:bold; margin-bottom:4px;'>请选择最远整理日期</div>",
            unsafe_allow_html=True,
        )
        since = st.date_input(
            "请选择最远整理日期",
            value=st.session_state.parse_since_date,
            max_value=date.today(),
            label_visibility="collapsed",
            key="parse_since_date_picker",
            help="只整理修改时间晚于此日期的简历（含当天）",
            disabled=not is_idle,
        )
        st.session_state.parse_since_date = since

    # btn_cols[1] 是间隔占位列，不放内容

    # 按钮列加顶部占位，与日期输入框底部对齐
    for _i in (2, 3, 4, 5):
        btn_cols[_i].markdown("<div style='height:24px'></div>", unsafe_allow_html=True)

    if is_idle:
        start_btn = btn_cols[2].button(
            "🚀 自动解析入库",
            type="primary",
            disabled=(not ai_service.is_configured) or (not all_files),
            key="parse_start_btn",
        )
        stop_btn = False
    else:
        start_btn = False
        stop_btn = btn_cols[2].button(
            "⏹ 停止解析任务",
            type="secondary",
            disabled=is_stopping,
            key="parse_stop_btn",
        )

    refresh_btn = btn_cols[3].button("🔄 重新扫描", disabled=not is_idle, key="parse_refresh_btn")
    clear_btn = btn_cols[4].button("🧹 清空窗口", disabled=not is_idle, key="parse_clear_btn")

    # 运行中显示进度副文本（也只在窗口外的"状态条"，不输出任务信息）
    if is_running or is_stopping:
        done = st.session_state.parse_index
        total_q = len(st.session_state.parse_queue)
        btn_cols[5].progress(
            done / max(total_q, 1),
            text=f"进度 {done}/{total_q}" + ("（正在停止…）" if is_stopping else ""),
        )

    # ---------- 按钮事件处理 ----------
    if clear_btn:
        st.session_state.parse_log_lines = []
        append_scan_summary(st.session_state.parse_log_lines, all_files)
        st.rerun()

    if refresh_btn:
        append_scan_summary(st.session_state.parse_log_lines, all_files)
        st.rerun()

    if start_btn:
        # 任务开始：按日期过滤文件列表 + 初始化统计
        st.session_state.parse_task_state = "running"
        since_dt = datetime.combine(st.session_state.parse_since_date, datetime.min.time())
        filtered_files = [
            f for f in all_files
            if datetime.fromtimestamp(f["path"].stat().st_mtime) >= since_dt
        ]
        filtered_files, pre_skip_files = filter_duplicate_files_before_ai(filtered_files)
        st.session_state.parse_queue = filtered_files
        st.session_state.parse_index = 0
        st.session_state.parse_results = {
            "success_count": 0,
            "skip_count": len(pre_skip_files),
            "fail_count": 0,
            "success_files": [],
            "skip_files": pre_skip_files,
            "failed_files": {
                "空白/损坏": [],
                "文本过短": [],
                "AI 解析异常": [],
                "AI 返回空结果": [],
                "入库失败": [],
            },
        }
        ts = datetime.now().strftime("%H:%M:%S")
        st.session_state.parse_log_lines.append("─" * 60)
        st.session_state.parse_log_lines.append(f"[{ts}] 🚀 解析任务开始")
        st.session_state.parse_log_lines.append(f"  日期过滤：≥ {st.session_state.parse_since_date.isoformat()}")
        st.session_state.parse_log_lines.append(f"  AI 调用前预去重跳过：{len(pre_skip_files)} 份")
        st.session_state.parse_log_lines.append(f"  待 AI 处理简历：{len(filtered_files)} 份（原始扫描 {len(all_files)} 份）")
        st.session_state.parse_log_lines.append("─" * 60)
        st.rerun()

    if stop_btn:
        st.session_state.parse_task_state = "stopping"
        ts = datetime.now().strftime("%H:%M:%S")
        st.session_state.parse_log_lines.append(f"[{ts}] ⏹ 收到停止指令，正在收尾…")
        st.rerun()

    # ---------- 任务执行主循环（每次 rerun 处理 1 份，然后再 rerun）----------
    if is_running or is_stopping:
        queue = st.session_state.parse_queue
        idx = st.session_state.parse_index
        results = st.session_state.parse_results

        def log(line: str) -> None:
            st.session_state.parse_log_lines.append(line)

        # 停止 or 队列处理完 → 写收尾统计并归位
        if is_stopping or idx >= len(queue):
            end_ts = datetime.now().strftime("%H:%M:%S")
            log("─" * 60)
            log(f"[{end_ts}] 🏁 解析任务{'已停止' if is_stopping else '结束'}")
            log("")
            log("📊 任务统计")
            log(f"  • 计划处理：{len(queue)} 份")
            log(f"  • 实际处理：{idx} 份")
            log(f"  • ✅ 成功入库：{results['success_count']} 份")
            log(f"  • 🔄 重复跳过：{results['skip_count']} 份")
            log(f"  • ❌ 解析失败：{results['fail_count']} 份")
            if is_stopping:
                log(f"  • ⏹ 未处理（被停止）：{len(queue) - idx} 份")

            if results["fail_count"] > 0:
                log("")
                log("❌ 失败明细（按原因分组）")
                for reason, items in results["failed_files"].items():
                    if not items:
                        continue
                    log(f"  【{reason}】{len(items)} 份")
                    for item in items:
                        log(f"    - {item}")

            if results["skip_count"] > 0:
                log("")
                log("🔄 重复跳过明细")
                for item in results["skip_files"]:
                    log(f"  - {item}")

            if results["success_count"] > 0:
                log("")
                log("✅ 成功入库明细")
                for item in results["success_files"]:
                    log(f"  - {item}")

            log("")
            log("💡 顶部 4 个 metric 数字将在下次刷新页面（F5 或切 tab）后更新。")

            # 归位
            st.session_state.parse_task_state = "idle"
            st.session_state.parse_queue = []
            st.session_state.parse_index = 0
            st.rerun()

        # 正常处理 1 份
        file_info = queue[idx]
        path = file_info["path"]
        platform = file_info["platform"]
        fname = file_info["name"]

        log(f"[{idx+1}/{len(queue)}] 🔍 处理：{fname}（{platform}）")

        # 1) 空白/损坏检测
        if should_skip_empty_or_corrupted(path):
            log("           ⚠️ 跳过 — 文件空白或损坏")
            results["failed_files"]["空白/损坏"].append(fname)
            results["fail_count"] += 1
        else:
            # 2) 文本提取
            raw_text = extract_text(path)
            # PDF 文本过短（< 200）通常是纯图像简历，尝试 PaddleOCR 回退
            if path.suffix.lower() == ".pdf" and len(raw_text.strip()) < 200:
                from recruitment_assistant.parsers.ocr_service import (
                    is_paddleocr_available,
                    ocr_pdf_to_text,
                )
                log(f"           🖼️ PDF 文本过短（{len(raw_text.strip())} 字符），疑似图像简历，启动 OCR 回退…")
                if not is_paddleocr_available():
                    log("           ⚠️ PaddleOCR 未安装（pip install paddlepaddle paddleocr），跳过 OCR")
                else:
                    try:
                        ocr_text = ocr_pdf_to_text(path, log=log)
                        if len(ocr_text.strip()) > len(raw_text.strip()):
                            log(f"           ✅ OCR 完成，得到 {len(ocr_text)} 字符")
                            raw_text = ocr_text
                        else:
                            log("           ⚠️ OCR 未识别出更多文本")
                    except Exception as exc:
                        log(f"           ❌ OCR 异常：{exc}")
            if len(raw_text.strip()) < 50:
                log(f"           ⚠️ 跳过 — 提取文本过短（{len(raw_text)} 字符）")
                results["failed_files"]["文本过短"].append(fname)
                results["fail_count"] += 1
            else:
                if has_garbled_text(raw_text):
                    log("           ⚠️ PDF 含解析乱码（嵌入子集字体缺 ToUnicode 映射），邮箱/手机/年龄等数字字段可能识别失败，请人工补录。")
                log(f"           📄 文本提取成功（{len(raw_text)} 字符），调用 AI 结构化…")
                # 3) AI 结构化
                candidate_data = None
                ai_failed = False
                try:
                    candidate_data = ai_service.parse_resume_text(raw_text, source_name=fname)
                except Exception as exc:
                    # pydantic ValidationError 是多行的，首行只是计数（"1 validation error for ..."）,
                    # 后续行才有 字段路径 / 实际值 / 错误类型，必须一并打到日志里才能定位 bug
                    err_lines = str(exc).splitlines() or [repr(exc)]
                    log(f"           ❌ AI 解析异常：{err_lines[0]}")
                    for sub in err_lines[1:8]:
                        log(f"              {sub}")
                    results["failed_files"]["AI 解析异常"].append(f"{fname}（{err_lines[0]}）")
                    results["fail_count"] += 1
                    ai_failed = True

                if not ai_failed and not candidate_data:
                    log("           ⚠️ AI 返回空结果")
                    results["failed_files"]["AI 返回空结果"].append(fname)
                    results["fail_count"] += 1
                elif candidate_data:
                    fname_name, fname_age, fname_edu = infer_candidate_from_filename(fname)
                    filled = []
                    if not candidate_data.age and fname_age:
                        candidate_data.age = fname_age
                        filled.append(f"年龄={fname_age}")
                    if not candidate_data.education_level and fname_edu:
                        candidate_data.education_level = fname_edu
                        filled.append(f"学历={fname_edu}")
                    if (not candidate_data.name or candidate_data.name in ("未知", "未识别")) and fname_name:
                        candidate_data.name = fname_name
                        filled.append(f"姓名={fname_name}")
                    if filled:
                        log(f"           🪪 从文件名补全字段：{' / '.join(filled)}")
                    log(
                        f"           🤖 AI 识别 — 姓名:{candidate_data.name} "
                        f"电话:{candidate_data.phone or '-'} "
                        f"学历:{candidate_data.education_level or '-'} "
                        f"城市:{candidate_data.current_city or '-'}"
                    )
                    # 4)+5)+6) 去重 → 补充来源 → 入库
                    session = create_resume_session()
                    try:
                        svc = ResumeArchiveService(session)
                        if svc.is_duplicate(
                            phone=candidate_data.phone,
                            name=candidate_data.name,
                            age=candidate_data.age,
                            education_level=candidate_data.education_level,
                        ):
                            log("           🔄 去重跳过 — 候选人已存在于数据库")
                            results["skip_files"].append(f"{candidate_data.name}（{fname}）")
                            results["skip_count"] += 1
                        else:
                            candidate_data.resume_source = ResumeSourceCreate(
                                source_platform=normalize_platform(platform),
                                file_name=fname,
                                file_type=path.suffix.lstrip(".").upper(),
                                file_path=str(path),
                                attachment_works_path=find_attachment_works_path_for(path),
                                crawl_time=datetime.now(),
                            )
                            try:
                                created = svc.create_candidate(candidate_data)
                                log(f"           ✅ 入库成功 — candidate_id={created.candidate_id}")
                                results["success_files"].append(
                                    f"{candidate_data.name}（{platform} - {fname}）"
                                )
                                results["success_count"] += 1
                            except Exception as exc:
                                log(f"           ❌ 入库失败：{exc}")
                                results["failed_files"]["入库失败"].append(f"{fname}（{exc}）")
                                results["fail_count"] += 1
                    finally:
                        session.close()

        # 处理完一份，索引 +1，触发下一轮 rerun
        st.session_state.parse_index = idx + 1
        st.rerun()


# ==================== Tab 2: 简历库浏览 ====================
def _fmt_date(d) -> str:
    """date 或 datetime 对象 → YYYY-MM-DD；None → ''。"""
    if not d:
        return ""
    try:
        return d.strftime("%Y-%m-%d")
    except Exception:
        return str(d)


@st.dialog("发起面试邀约")
def _open_invite_dialog():
    """弹窗：选岗位（可不选）→ 检查去重 → 写库。

    用 session_state['invite_dialog_cid'] 在浏览页 → dialog 之间传候选人 ID。
    """
    cid = st.session_state.get("invite_dialog_cid")
    if not cid:
        st.error("未选中候选人")
        return
    session = create_resume_session()
    svc = ResumeArchiveService(session)
    try:
        cand = svc.get_candidate(cid)
        if not cand:
            st.error("候选人不存在")
            return

        st.markdown(
            f"**候选人：** {cand.name}（"
            f"{cand.age or '?'}岁 · {cand.education_level or '-'}"
            f"）"
        )

        # 去重检测：同一候选人同时只能有 1 条 pending
        already_pending = svc.has_pending_invitation(cid)
        if already_pending:
            st.warning("⚠️ 该候选人已有进行中的邀约，请先到「面试管理」页面取消或完成。")

        # 岗位下拉（可不选）— 从 PostgreSQL 获取
        pg_sess = create_session()
        invite_job_svc = JobService(pg_sess)
        positions = invite_job_svc.list_positions(status="active")
        pg_sess.close()
        pos_options = ["（不指定岗位）"] + [
            f"{p.position_id} | {p.title}（{p.department or '-'} · {p.work_city or '-'}）"
            for p in positions
        ]
        pick = st.selectbox(
            "拟招聘岗位（可选）",
            pos_options,
            key=f"invite_pos_pick_{cid}",
        )
        notes = st.text_area(
            "备注（可选）",
            key=f"invite_notes_{cid}",
            height=80,
        )

        confirm_col, cancel_col = st.columns(2)
        if confirm_col.button(
            "✅ 确认邀约",
            type="primary",
            disabled=already_pending,
            use_container_width=True,
            key=f"invite_confirm_{cid}",
        ):
            position_id = None
            if pick != "（不指定岗位）":
                position_id = int(pick.split(" | ", 1)[0])
            svc.create_invitation(candidate_id=cid, position_id=position_id, notes=notes)
            st.session_state.pop("invite_dialog_cid", None)
            st.rerun()
        if cancel_col.button(
            "取消",
            use_container_width=True,
            key=f"invite_cancel_{cid}",
        ):
            st.session_state.pop("invite_dialog_cid", None)
            st.rerun()
    finally:
        session.close()


def _render_candidate_detail(c, svc, session) -> None:
    """右侧详情区：把候选人的所有字段按段落渲染。"""
    head_cols = st.columns([4, 1, 1, 1])
    head_cols[0].markdown(f"## {c.name}")
    # ⭐ 关注复选框（实时落库）
    new_fav = head_cols[1].checkbox(
        "⭐ 关注",
        value=bool(c.is_favorite),
        key=f"browse_fav_{c.candidate_id}",
    )
    if int(new_fav) != int(c.is_favorite or 0):
        svc.update_candidate_field(c.candidate_id, is_favorite=int(new_fav))
        st.rerun()
    if head_cols[2].button("📧 面试邀约", key=f"browse_invite_{c.candidate_id}"):
        st.session_state["invite_dialog_cid"] = c.candidate_id
        _open_invite_dialog()
    if head_cols[3].button("🗑️ 删除", key=f"browse_del_{c.candidate_id}"):
        svc.delete_candidate(c.candidate_id)
        st.session_state.pop("browse_selected_cid", None)
        session.close()
        st.rerun()

    # 基本信息
    st.markdown("##### 基本信息")
    info_items = [
        ("性别", c.gender),
        ("年龄", f"{c.age}岁" if c.age else None),
        ("生日", _fmt_date(c.birth_date) or None),
        ("学历", c.education_level),
        ("现居城市", c.current_city),
        ("手机", c.phone),
        ("邮箱", c.email),
        ("微信", c.wechat),
    ]
    info_lines = [f"- **{k}**：{v}" for k, v in info_items if v]
    st.markdown("\n".join(info_lines) if info_lines else "_（无）_")

    # 教育经历
    st.markdown("##### 教育经历")
    if c.educations:
        for edu in c.educations:
            head = f"- **{edu.school_name}**"
            if edu.start_date or edu.end_date:
                head += f"（{_fmt_date(edu.start_date) or '?'} — {_fmt_date(edu.end_date) or '至今'}）"
            st.markdown(head)
            subs = [edu.education_level, edu.degree, edu.major,
                    "全日制" if edu.is_full_time else "非全日制"]
            subs = [s for s in subs if s]
            if subs:
                st.markdown(f"  - {' · '.join(subs)}")
    else:
        st.markdown("_（无）_")

    # 工作经历
    st.markdown("##### 工作经历")
    if c.work_experiences:
        for w in c.work_experiences:
            head = f"- **{w.company_name}**"
            if w.start_date or w.end_date:
                head += f"（{_fmt_date(w.start_date) or '?'} — {_fmt_date(w.end_date) or '至今'}）"
            st.markdown(head)
            subs = [s for s in (w.position, w.industry) if s]
            if subs:
                st.markdown(f"  - {' · '.join(subs)}")
            if w.job_content:
                st.markdown(f"  - 工作内容：{w.job_content}")
    else:
        st.markdown("_（无）_")

    # 项目经历
    st.markdown("##### 项目经历")
    if c.project_experiences:
        for p in c.project_experiences:
            head = f"- **{p.project_name}**"
            if p.project_date:
                head += f"（{p.project_date}）"
            st.markdown(head)
            if p.project_role:
                st.markdown(f"  - 角色：{p.project_role}")
            if p.project_desc:
                st.markdown(f"  - 项目描述：{p.project_desc}")
            if p.project_duty:
                st.markdown(f"  - 我的职责：{p.project_duty}")
            if p.project_result:
                st.markdown(f"  - 项目成果：{p.project_result}")
    else:
        st.markdown("_（无）_")

    # 技能 / 证书（按 skill_type 分组）
    st.markdown("##### 技能 / 证书")
    if c.skills:
        by_type: dict[str, list] = defaultdict(list)
        for s in c.skills:
            by_type[s.skill_type or "其他"].append(s)
        for stype, items in by_type.items():
            parts = []
            for s in items:
                name = s.skill_name or ""
                if s.proficiency:
                    name += f"（{s.proficiency}）"
                parts.append(name)
            st.markdown(f"- **{stype}**：{'、'.join(parts)}")
    else:
        st.markdown("_（无）_")

    # 求职意向
    st.markdown("##### 求职意向")
    if c.job_intention:
        ji = c.job_intention
        ji_items = [
            ("目标岗位", ji.target_position),
            ("期望城市", ji.target_city),
            ("期望薪资", ji.expected_salary),
            ("求职状态", ji.job_status),
        ]
        ji_lines = [f"- **{k}**：{v}" for k, v in ji_items if v]
        st.markdown("\n".join(ji_lines) if ji_lines else "_（无）_")
    else:
        st.markdown("_（无）_")

    # 荣誉
    st.markdown("##### 荣誉")
    if c.honors:
        for h in c.honors:
            line = f"- **{h.honor_name}**"
            extras = [x for x in (h.honor_level, _fmt_date(h.honor_date)) if x]
            if extras:
                line += f"（{' · '.join(extras)}）"
            st.markdown(line)
    else:
        st.markdown("_（无）_")

    # 自我评价（完整展示，不截断）
    st.markdown("##### 自我评价")
    st.markdown(c.self_intro if c.self_intro else "_（无）_")

    # 简历来源（含本地文件地址）
    st.markdown("##### 简历来源")
    if c.resume_source:
        rs = c.resume_source
        src_lines = []
        if rs.source_platform:
            src_lines.append(f"- **来源平台**：{rs.source_platform}")
        if rs.file_name:
            src_lines.append(f"- **原始文件名**：{rs.file_name}")
        if rs.file_type:
            src_lines.append(f"- **文件类型**：{rs.file_type}")
        if rs.crawl_time:
            src_lines.append(f"- **入库时间**：{rs.crawl_time.strftime('%Y-%m-%d %H:%M:%S')}")
        st.markdown("\n".join(src_lines) if src_lines else "_（无）_")
        if rs.file_path:
            st.markdown("**📎 简历文件本地地址：**")
            st.code(rs.file_path, language=None)
            fp = Path(rs.file_path)
            file_exists = fp.exists()
            dir_exists = fp.parent.exists()
            op_cols = st.columns([1, 1, 6])
            if op_cols[0].button("📄 打开", key=f"browse_open_file_{c.candidate_id}",
                                  disabled=not file_exists,
                                  help="用默认应用打开简历文件"):
                try:
                    os.startfile(str(fp))
                except OSError as exc:
                    st.error(f"打开失败：{exc}")
            if op_cols[1].button("📁 访问目录", key=f"browse_open_dir_{c.candidate_id}",
                                  disabled=not dir_exists,
                                  help="在资源管理器中打开文件所在目录"):
                try:
                    os.startfile(str(fp.parent))
                except OSError as exc:
                    st.error(f"打开目录失败：{exc}")
            if not file_exists:
                st.warning("⚠️ 文件不存在，可能已被移动或删除")
        if getattr(rs, "attachment_works_path", None):
            st.markdown("**🎨 附件作品文件地址：**")
            st.code(rs.attachment_works_path, language=None)
            wp = Path(rs.attachment_works_path)
            works_exists = wp.exists()
            works_cols = st.columns([1, 7])
            if works_cols[0].button("📄 打开", key=f"browse_open_works_{c.candidate_id}",
                                     disabled=not works_exists,
                                     help="用默认应用打开附件作品文件"):
                try:
                    os.startfile(str(wp))
                except OSError as exc:
                    st.error(f"打开失败：{exc}")
            if not works_exists:
                st.warning("⚠️ 附件作品文件不存在，可能已被移动或删除")
    else:
        st.markdown("_（无）_")


with tabs[1]:
    st.markdown("### 简历库浏览")

    # 过滤条
    filt_cols = st.columns([1.4, 1.0, 1.0, 1.0, 1.0])
    f_name = filt_cols[0].text_input("姓名", key="browse_name")
    f_city = filt_cols[1].text_input("城市", key="browse_city")
    f_edu = filt_cols[2].selectbox(
        "学历", ["全部", "博士", "硕士", "本科", "大专", "高中", "中专"], key="browse_edu"
    )
    f_platform = filt_cols[3].selectbox(
        "来源", ["全部", "BOSS直聘", "智联招聘", "51前程无忧"], key="browse_platform"
    )
    f_mark = filt_cols[4].selectbox("标记", ["全部", "关注"], key="browse_mark")

    session = create_resume_session()
    svc = ResumeArchiveService(session)
    # 一次拉全部（不分页，左侧列表滚动浏览）。page_size 设大可保证全量返回。
    candidates, total = svc.list_candidates(
        page=1,
        page_size=10000,
        name=f_name or None,
        city=f_city or None,
        education_level=f_edu if f_edu != "全部" else None,
        platform=f_platform if f_platform != "全部" else None,
        favorite_only=(f_mark == "关注"),
    )
    st.caption(f"共 {total} 条候选人")

    if not candidates:
        st.info("暂无候选人数据。请先在「简历自动解析入库」中导入简历。")
        session.close()
    else:
        # session_state 记忆当前选中候选人。过滤后若原选中不在结果集内，回退第一条
        valid_cids = {c.candidate_id for c in candidates}
        if st.session_state.get("browse_selected_cid") not in valid_cids:
            st.session_state.browse_selected_cid = candidates[0].candidate_id

        list_col, detail_col = st.columns([1, 3])

        with list_col:
            st.markdown("**候选人列表**")
            # 列表内按钮基础样式：白底黑字 + 左对齐 + primary 态高亮
            # 注：关注高亮改用 label "⭐ " 前缀（见 for 循环），不再依赖 nth-child CSS
            # —— 早期版本用 nth-child 选择器在 Streamlit 1.56 的 DOM 嵌套层数下命中失败，
            #    且条件性注入 fav CSS 会让整个容器位置漂移
            st.markdown(
                """
                <style>
                div[data-testid="stVerticalBlockBorderWrapper"] button {
                    text-align: left !important;
                    justify-content: flex-start !important;
                    background: var(--color-surface) !important;
                    color: var(--color-text) !important;
                    border: 1px solid var(--color-border) !important;
                    font-weight: normal !important;
                }
                div[data-testid="stVerticalBlockBorderWrapper"] button p {
                    text-align: left !important;
                    color: var(--color-text) !important;
                }
                div[data-testid="stVerticalBlockBorderWrapper"] button:hover {
                    background: var(--color-hover) !important;
                    border-color: var(--color-border-strong, var(--color-border)) !important;
                }
                div[data-testid="stVerticalBlockBorderWrapper"] button[kind="primary"] {
                    background: var(--color-primary-soft) !important;
                    border-color: var(--color-primary) !important;
                    font-weight: 600 !important;
                }
                </style>
                """,
                unsafe_allow_html=True,
            )
            with st.container(height=1400, border=True):
                for c in candidates:
                    is_selected = c.candidate_id == st.session_state.browse_selected_cid
                    prefix = "⭐ " if c.is_favorite else ""
                    label = f"{prefix}{c.name} | {c.age or '?'}岁 | {c.education_level or '-'}"
                    if st.button(
                        label,
                        key=f"browse_pick_{c.candidate_id}",
                        use_container_width=True,
                        type="primary" if is_selected else "secondary",
                    ):
                        st.session_state.browse_selected_cid = c.candidate_id
                        st.rerun()

        with detail_col:
            selected = next(
                (c for c in candidates if c.candidate_id == st.session_state.browse_selected_cid),
                None,
            )
            if selected:
                _render_candidate_detail(selected, svc, session)

        session.close()

# ==================== Tab 3: 招聘岗位录入/匹配 ====================
SALARY_LOW_OPTIONS = ["不限", "3K", "5K", "8K", "10K", "12K", "15K", "20K", "25K", "30K", "40K", "50K"]
SALARY_HIGH_OPTIONS = ["不限", "5K", "8K", "10K", "12K", "15K", "20K", "25K", "30K", "40K", "50K", "80K", "100K"]
EDU_OPTIONS = ["不限", "大专", "本科", "硕士以上"]
EXP_OPTIONS = ["不限", "1-3年", "3-5年", "5-10年", "10年以上"]


# ---------- 岗位匹配工具函数 ----------

def _build_candidate_dicts(all_candidates) -> list[dict]:
    """将 ORM Candidate 列表转换为 AI 匹配所需的字典列表。"""
    candidate_dicts = []
    for c in all_candidates:
        skills_str = "、".join(s.skill_name or "" for s in c.skills if s.skill_name)[:100]

        core_skills = [s.skill_name for s in c.skills if s.is_core and s.skill_name][:8]
        core_skills_str = "、".join(core_skills) if core_skills else "-"

        work_str = "; ".join(
            f"{w.company_name}({w.position or '-'})"
            for w in (c.work_experiences or [])[:3]
        )

        projects = c.project_experiences[:2] if c.project_experiences else []
        projects_str = "; ".join(
            f"{p.project_name}（{p.project_role or '参与'}）"
            for p in projects if p.project_name
        ) or "-"

        honors = [h.honor_name for h in (c.honors or []) if h.honor_name][:3]
        honors_str = "、".join(honors) if honors else "-"

        years_exp = "-"
        if c.work_experiences:
            from datetime import date as _date
            total_months = 0
            for w in c.work_experiences:
                if w.start_date:
                    end = w.end_date or _date.today()
                    months = (end.year - w.start_date.year) * 12 + (end.month - w.start_date.month)
                    total_months += max(0, months)
            if total_months > 0:
                years_exp = f"{round(total_months / 12, 1)}年"

        candidate_dicts.append({
            "candidate_id": c.candidate_id,
            "name": c.name,
            "education_level": c.education_level or "-",
            "current_city": c.current_city or "-",
            "position": c.work_experiences[0].position if c.work_experiences else "-",
            "skills": skills_str or "-",
            "core_skills": core_skills_str,
            "work_summary": work_str or "-",
            "years_of_experience": years_exp,
            "projects": projects_str,
            "honors": honors_str,
        })
    return candidate_dicts


def _build_full_jd(pos) -> str:
    """拼接岗位的完整 JD 文本（岗位职责 + 任职要求）。"""
    parts = []
    if pos.responsibilities:
        parts.append(f"【岗位职责】\n{pos.responsibilities}")
    if pos.job_requirements:
        parts.append(f"【任职要求】\n{pos.job_requirements}")
    return "\n\n".join(parts) if parts else pos.title


def _diagnose_fk_failure(debug_logger, svc, position_id: int, failed_cid: int):
    """首次保存失败时执行一次性 FK 诊断，结果写入调试日志。"""
    from sqlalchemy import text
    diag = {}
    try:
        sess = svc.session
        diag["pragma_foreign_keys"] = sess.execute(text("PRAGMA foreign_keys")).scalar()

        row = sess.execute(
            text("SELECT candidate_id FROM candidates WHERE candidate_id = :cid"),
            {"cid": failed_cid},
        ).fetchone()
        diag["candidate_exists_raw"] = row is not None

        fk_issues = sess.execute(text("PRAGMA foreign_key_check(position_matches)")).fetchall()
        diag["fk_check_issues"] = [list(r) for r in fk_issues[:20]] if fk_issues else []

        tbl_info = sess.execute(text("SELECT sql FROM sqlite_master WHERE name='position_matches'")).scalar()
        diag["table_ddl"] = tbl_info

        cnt = sess.execute(text("SELECT COUNT(*) FROM candidates")).scalar()
        diag["candidates_total"] = cnt

        id_range = sess.execute(
            text("SELECT MIN(candidate_id), MAX(candidate_id) FROM candidates")
        ).fetchone()
        diag["candidates_id_range"] = list(id_range) if id_range else []

        sample_around = sess.execute(
            text("SELECT candidate_id FROM candidates WHERE candidate_id BETWEEN :lo AND :hi ORDER BY candidate_id"),
            {"lo": failed_cid - 5, "hi": failed_cid + 5},
        ).fetchall()
        diag["candidates_near_failed_id"] = [r[0] for r in sample_around]

    except Exception as exc:
        diag["diag_error"] = str(exc)

    debug_logger.log("fk_diagnosis", f"首次保存失败的 FK 诊断 (candidate_id={failed_cid})", diag)


def _run_single_position_match(pos, svc, ai_service, candidate_dicts) -> tuple[int, int, bool]:
    """对单个岗位执行 AI 匹配（串行批次），返回 (save_ok, save_fail, timed_out)。"""
    from recruitment_assistant.utils.match_debug_logger import MatchDebugLogger
    from sqlalchemy import text

    POSITION_MATCH_TIMEOUT = 180

    # ✨ 创建独立的调试日志记录器
    logger.info("[调试日志] 开始为岗位 {} (id={}) 创建调试日志", pos.title, pos.position_id)
    debug_logger = MatchDebugLogger(pos.position_id, pos.title)
    logger.info("[调试日志] MatchDebugLogger 已创建")

    try:
        svc.clear_position_matches(pos.position_id)
        total = len(candidate_dicts)
        if total == 0:
            debug_logger.log("empty", "候选人列表为空，跳过匹配")
            return 0, 0, False

        # ✨ 验证数据一致性：获取数据库中实际存在的候选人ID集合
        resume_session = create_resume_session()
        try:
            result = resume_session.execute(text("SELECT candidate_id FROM candidates"))
            valid_ids = {row[0] for row in result.fetchall()}
            logger.info("[数据验证] 数据库中实际有 {} 个候选人", len(valid_ids))
            debug_logger.log("data_validation", f"数据库实际候选人数量: {len(valid_ids)}", {
                "db_count": len(valid_ids),
                "orm_count": total,
                "id_sample": list(valid_ids)[:20] if valid_ids else []
            })

            # 过滤掉不存在的候选人
            candidate_dicts = [c for c in candidate_dicts if c.get("candidate_id") in valid_ids]
            filtered_count = len(candidate_dicts)
            if filtered_count < total:
                logger.warning("[数据验证] 过滤掉 {} 个不存在的候选人，剩余 {}",
                              total - filtered_count, filtered_count)
                debug_logger.log("data_filter", f"过滤掉不存在的候选人", {
                    "original_count": total,
                    "filtered_count": filtered_count,
                    "removed_count": total - filtered_count
                })

            if filtered_count == 0:
                debug_logger.log("empty_after_filter", "过滤后无有效候选人")
                return 0, 0, False

            total = filtered_count
        finally:
            resume_session.close()

        chunk_size = max(3, min(20, total // 5))
        full_jd = _build_full_jd(pos)

        # ✨ 记录候选人数据
        resume_session = create_resume_session()
        try:
            resume_svc = ResumeArchiveService(resume_session)
            candidates_orm, _ = resume_svc.list_candidates(page=1, page_size=9999)
            debug_logger.log_candidates(candidates_orm, candidate_dicts)
        finally:
            resume_session.close()

        chunks = [candidate_dicts[i:i + chunk_size] for i in range(0, total, chunk_size)]
        all_results = []
        start_time = _time.monotonic()
        timed_out = False
        for chunk_idx, chunk in enumerate(chunks):
            elapsed = _time.monotonic() - start_time
            if elapsed > POSITION_MATCH_TIMEOUT:
                logger.warning("[岗位匹配] 岗位 '{}' 超时(已用{:.0f}s>{}s)，已完成 {}/{} 批次",
                               pos.title, elapsed, POSITION_MATCH_TIMEOUT, chunk_idx, len(chunks))
                debug_logger.log("timeout", f"岗位匹配超时，已完成 {chunk_idx}/{len(chunks)} 批次")
                timed_out = True
                break
            try:
                # ✨ 传递 debug_logger 给 AI 服务
                results = ai_service.match_candidates(full_jd, chunk, debug_logger=debug_logger)
                all_results.extend(results)
            except Exception as exc:
                logger.warning("[岗位匹配] 批次失败({}): {}", pos.title, exc)
                debug_logger.log_error(f"batch_{chunk_idx}", exc)

        # ✨ 记录保存尝试
        candidate_ids = [r.get("candidate_id") for r in all_results if r.get("candidate_id")]
        debug_logger.log_save_attempt(len(all_results), candidate_ids)

        save_ok, save_fail = 0, 0
        failed_ids = []
        fk_diagnosed = False
        for r in all_results:
            cid = r.get("candidate_id")
            score = r.get("match_score", 0)
            reason = r.get("reason", "")
            dimensions = r.get("dimensions")
            if cid and isinstance(score, (int, float)):
                try:
                    svc.save_position_match(
                        pos.position_id, int(cid), int(score), reason,
                        dimensions=dimensions,
                    )
                    save_ok += 1
                except Exception as exc:
                    save_fail += 1
                    failed_ids.append(int(cid))
                    debug_logger.log_error(f"save_candidate_{cid}", exc)

                    if not fk_diagnosed:
                        fk_diagnosed = True
                        _diagnose_fk_failure(debug_logger, svc, pos.position_id, int(cid))

        # ✨ 记录保存结果
        debug_logger.log_save_result(save_ok, save_fail, failed_ids)

        return save_ok, save_fail, timed_out

    finally:
        # ✨ 完成日志记录并保存
        log_path = debug_logger.finalize()
        logger.info("[调试日志] 匹配调试日志已保存: {}", log_path)


# ---------- Excel 批量导入工具函数 ----------

def _parse_jd_sections(jd_text: str) -> tuple[str, str]:
    """将 JD 文本按「岗位职责」「任职要求」关键词拆分为两段。

    返回 (responsibilities, job_requirements)。
    若未找到分隔标记，整段文本归入 job_requirements。
    """
    if not jd_text or not jd_text.strip():
        return "", ""

    # 常见的任职要求段落标题模式
    req_pattern = re.compile(
        r"\n\s*(任职要求|任职资格|岗位要求|招聘要求|职位要求)\s*[:：]?\s*\n",
        re.IGNORECASE,
    )
    match = req_pattern.search(jd_text)
    if match:
        resp_part = jd_text[:match.start()].strip()
        req_part = jd_text[match.end():].strip()
        # 去掉岗位职责段落自身的标题行
        resp_part = re.sub(
            r"^(岗位职责|工作职责|职位职责|主要职责)\s*[:：]?\s*\n?",
            "", resp_part, count=1, flags=re.IGNORECASE,
        ).strip()
        return resp_part, req_part

    # 无法拆分时，整段归入 job_requirements
    return "", jd_text.strip()


def _format_salary_wan(low, high) -> str:
    """将万元数值对 (1.5, 3) 转为 UI 格式字符串 '15K-30K'。"""
    def _wan_to_k(val) -> str | None:
        if val is None:
            return None
        try:
            k = int(float(val) * 10)
            return f"{k}K"
        except (ValueError, TypeError):
            return None

    low_k = _wan_to_k(low)
    high_k = _wan_to_k(high)
    if low_k and high_k:
        return f"{low_k}-{high_k}"
    if low_k:
        return f"≥{low_k}"
    if high_k:
        return f"≤{high_k}"
    return ""


def _format_experience(val) -> str | None:
    """将工作年限值规范化为字符串。"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return f"{int(val)}年"
    s = str(val).strip()
    return s if s else None


def _parse_excel_positions(file_bytes: bytes, file_name: str) -> tuple[list[dict], list[str]]:
    """解析上传的 xlsx 文件，返回 (positions_list, errors_list)。

    每个 position 是可直接传给 JobService.create_position() 的关键字字典。
    """
    import openpyxl

    errors: list[str] = []
    positions: list[dict] = []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    except Exception as exc:
        return [], [f"无法打开 Excel 文件：{exc}"]

    ws = wb.active
    if ws is None:
        wb.close()
        return [], ["Excel 文件中没有工作表"]

    # 读取表头并建立列名→索引映射
    header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
    if not header_row:
        wb.close()
        return [], ["Excel 文件第一行为空（缺少表头）"]

    headers = [str(cell.value or "").strip() for cell in header_row]

    # 查找必要列（岗位名称）
    col_map: dict[str, int] = {}
    KNOWN_HEADERS = {
        "岗位名称": "title",
        "部门": "department",
        "JD": "jd",
        "薪资下限": "salary_low",
        "薪资上限": "salary_high",
        "学历要求": "education",
        "工作年限": "experience",
        "工作经验": "experience",
    }
    # 处理可能存在的重复列名（如两个"薪资范围"列）
    salary_range_count = 0
    for idx, h in enumerate(headers):
        if h in KNOWN_HEADERS:
            col_map[KNOWN_HEADERS[h]] = idx
        elif h == "薪资范围":
            salary_range_count += 1
            if salary_range_count == 1:
                col_map["salary_low"] = idx
            elif salary_range_count == 2:
                col_map["salary_high"] = idx

    if "title" not in col_map:
        wb.close()
        return [], [f"未找到「岗位名称」列。当前表头：{headers}"]

    # 逐行解析
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        cells = [cell.value for cell in row]
        title_val = cells[col_map["title"]] if col_map.get("title") is not None else None
        if not title_val or not str(title_val).strip():
            continue  # 跳过空行

        jd_text = str(cells[col_map["jd"]]) if col_map.get("jd") is not None and cells[col_map["jd"]] else ""
        resp, req = _parse_jd_sections(jd_text)

        sal_low = cells[col_map["salary_low"]] if col_map.get("salary_low") is not None else None
        sal_high = cells[col_map["salary_high"]] if col_map.get("salary_high") is not None else None

        dept = str(cells[col_map["department"]]).strip() if col_map.get("department") is not None and cells[col_map["department"]] else ""
        edu = str(cells[col_map["education"]]).strip() if col_map.get("education") is not None and cells[col_map["education"]] else None
        exp_val = cells[col_map["experience"]] if col_map.get("experience") is not None else None

        positions.append({
            "title": str(title_val).strip(),
            "department": dept,
            "responsibilities": resp,
            "job_requirements": req,
            "salary_range": _format_salary_wan(sal_low, sal_high),
            "min_education": edu,
            "min_experience": _format_experience(exp_val),
        })

    wb.close()
    return positions, errors


def _render_position_form(prefill=None, key_suffix: str = "new"):
    """录入/编辑岗位表单。prefill 为 JobPosition 实例则做编辑预填，否则空表单。
    返回 (clicked, dict) — clicked=True 表示用户点了保存按钮。"""
    title = st.text_input("岗位名称", value=getattr(prefill, "title", "") or "",
                          key=f"posf_title_{key_suffix}")
    dept = st.text_input("部门", value=getattr(prefill, "department", "") or "",
                         key=f"posf_dept_{key_suffix}")
    responsibilities = st.text_area("岗位职责", value=getattr(prefill, "responsibilities", "") or "",
                                    height=200, key=f"posf_resp_{key_suffix}")
    job_req = st.text_area("任职要求", value=getattr(prefill, "job_requirements", "") or "",
                           height=200, key=f"posf_jobreq_{key_suffix}")

    sal_cols = st.columns([1, 0.2, 1])

    def _safe_index(opts, value, default=0):
        try:
            return opts.index(value)
        except ValueError:
            return default

    cur_low, cur_high = "不限", "不限"
    if prefill and prefill.salary_range and "-" in prefill.salary_range:
        parts = prefill.salary_range.split("-", 1)
        cur_low, cur_high = parts[0].strip(), parts[1].strip()
    sal_low = sal_cols[0].selectbox("薪资下限", SALARY_LOW_OPTIONS,
                                     index=_safe_index(SALARY_LOW_OPTIONS, cur_low),
                                     key=f"posf_sallow_{key_suffix}")
    sal_cols[1].markdown("<div style='text-align:center; padding-top:32px;'>—</div>",
                          unsafe_allow_html=True)
    sal_high = sal_cols[2].selectbox("薪资上限", SALARY_HIGH_OPTIONS,
                                      index=_safe_index(SALARY_HIGH_OPTIONS, cur_high),
                                      key=f"posf_salhigh_{key_suffix}")

    extra_cols = st.columns(2)
    edu = extra_cols[0].selectbox("学历要求", EDU_OPTIONS,
                                   index=_safe_index(EDU_OPTIONS, getattr(prefill, "min_education", None) or "不限"),
                                   key=f"posf_edu_{key_suffix}")
    exp = extra_cols[1].selectbox("工作年限", EXP_OPTIONS,
                                   index=_safe_index(EXP_OPTIONS, getattr(prefill, "min_experience", None) or "不限"),
                                   key=f"posf_exp_{key_suffix}")

    btn_label = "💾 保存修改" if prefill else "保存岗位"
    clicked = st.button(btn_label, type="primary", key=f"posf_btn_{key_suffix}")

    if sal_low == "不限" and sal_high == "不限":
        salary_range = ""
    elif sal_low == "不限":
        salary_range = f"≤{sal_high}"
    elif sal_high == "不限":
        salary_range = f"≥{sal_low}"
    else:
        salary_range = f"{sal_low}-{sal_high}"

    return clicked, {
        "title": title, "department": dept,
        "responsibilities": responsibilities, "job_requirements": job_req,
        "salary_range": salary_range,
        "min_education": None if edu == "不限" else edu,
        "min_experience": None if exp == "不限" else exp,
    }


@st.dialog("编辑岗位", width="large")
def _open_edit_position_dialog():
    pos_id = st.session_state.get("edit_pos_id")
    if not pos_id:
        st.error("缺少岗位 ID")
        return
    pg_session = create_session()
    job_svc = JobService(pg_session)
    pos = job_svc.get_by_id(pos_id)
    if not pos:
        pg_session.close()
        st.error("岗位不存在")
        return
    clicked, data = _render_position_form(prefill=pos, key_suffix=f"edit_{pos_id}")
    if clicked:
        if not data["title"]:
            st.warning("请填写岗位名称")
        else:
            job_svc.update_position(pos_id, **data)
            pg_session.close()
            st.session_state.pop("edit_pos_id", None)
            st.success("修改已保存")
            st.rerun()
    else:
        pg_session.close()


@st.dialog("确认清除")
def _confirm_clear_all_matches():
    st.warning("该操作会清除所有岗位的已匹配信息，是否确认清除?")
    cols = st.columns(2)
    if cols[0].button("确认清除", type="primary", use_container_width=True, key="confirm_clear_all"):
        session_tmp = create_resume_session()
        svc_tmp = ResumeArchiveService(session_tmp)
        for pos in st.session_state.get("_all_positions", []):
            svc_tmp.clear_position_matches(pos.position_id)
        session_tmp.close()
        logger.info("[清除匹配] 已清除所有岗位的匹配数据")
        st.rerun()
    if cols[1].button("取消", use_container_width=True, key="cancel_clear_all"):
        st.rerun()

with tabs[2]:
    st.markdown("### 招聘岗位录入/匹配")

    # ---- Excel 批量导入 ----
    with st.expander("📥 从 Excel 批量导入岗位", expanded=False):
        uploaded = st.file_uploader(
            "上传 .xlsx 文件",
            type=["xlsx"],
            key="pos_excel_upload",
            help="支持包含「岗位名称」列的 Excel 文件，可自动识别部门、JD、薪资、学历、工作年限等列",
        )
        if uploaded is not None:
            file_bytes = uploaded.getvalue()
            parsed_positions, parse_errors = _parse_excel_positions(file_bytes, uploaded.name)

            if parse_errors:
                for err in parse_errors:
                    st.error(err)
            elif not parsed_positions:
                st.warning("未从文件中解析到有效岗位数据。")
            else:
                # 预览表格
                preview_rows = []
                for p in parsed_positions:
                    resp_preview = (p["responsibilities"][:50] + "…") if len(p["responsibilities"]) > 50 else p["responsibilities"]
                    req_preview = (p["job_requirements"][:50] + "…") if len(p["job_requirements"]) > 50 else p["job_requirements"]
                    preview_rows.append({
                        "岗位名称": p["title"],
                        "部门": p["department"],
                        "薪资": p["salary_range"],
                        "学历": p["min_education"] or "",
                        "经验": p["min_experience"] or "",
                        "岗位职责": resp_preview,
                        "任职要求": req_preview,
                    })
                st.dataframe(preview_rows, use_container_width=True, hide_index=True)
                st.caption(f"共 {len(parsed_positions)} 条岗位，来源文件：{uploaded.name}")

                if st.button("✅ 确认导入", type="primary", key="pos_excel_import_btn"):
                    pg_sess_import = create_session()
                    import_svc = JobService(pg_sess_import)
                    success_count = 0
                    error_count = 0
                    for pos_data in parsed_positions:
                        try:
                            import_svc.create_position(
                                source_file_name=uploaded.name,
                                **pos_data,
                            )
                            success_count += 1
                        except Exception as exc:
                            error_count += 1
                            st.warning(f"导入「{pos_data['title']}」失败：{exc}")
                    pg_sess_import.close()
                    if success_count:
                        st.success(f"成功导入 {success_count} 条岗位！" + (f"（{error_count} 条失败）" if error_count else ""))
                    st.rerun()

    with st.expander("➕ 录入新岗位", expanded=False):
        clicked, data = _render_position_form(prefill=None, key_suffix="new")
        if clicked:
            if not data["title"]:
                st.warning("请填写岗位名称")
            else:
                pg_session = create_session()
                job_svc = JobService(pg_session)
                job_svc.create_position(**data)
                pg_session.close()
                st.success(f"岗位「{data['title']}」已保存")
                st.rerun()

    # ---- 两栏布局：左 1/3 岗位列表，右 2/3 匹配结果 ----
    pg_session = create_session()
    job_svc = JobService(pg_session)
    positions = job_svc.list_positions()

    if not positions:
        st.info("暂无岗位。请先录入招聘岗位。")
        pg_session.close()
    else:
        if "match_selected_pos" not in st.session_state or \
           st.session_state.match_selected_pos not in {p.position_id for p in positions}:
            st.session_state.match_selected_pos = positions[0].position_id

        pos_col, match_col = st.columns([1, 2])

        # ---- 左栏：岗位 expander 列表 ----
        with pos_col:
            st.markdown("**招聘岗位**")

            # ---- 批量匹配：多选岗位（最多5个）----
            is_running = st.session_state.get("auto_match_running", False)

            if not is_running:
                # 构建选项列表，处理重名岗位
                _seen_titles: dict[str, int] = {}
                _pos_labels: list[str] = []
                _label_to_id: dict[str, int] = {}
                for p in positions:
                    _seen_titles[p.title] = _seen_titles.get(p.title, 0) + 1
                _dup_titles = {t for t, c in _seen_titles.items() if c > 1}
                for p in positions:
                    label = f"{p.title} (ID:{p.position_id})" if p.title in _dup_titles else p.title
                    _pos_labels.append(label)
                    _label_to_id[label] = p.position_id

                selected_labels = st.multiselect(
                    "选择要匹配的岗位（最多5个）",
                    options=_pos_labels,
                    max_selections=5,
                    key="batch_match_selected",
                )
                selected_ids = [_label_to_id[lb] for lb in selected_labels]

            btn_cols = st.columns(2)
            with btn_cols[0]:
                if is_running:
                    if st.button("⏹ 停止自动匹配", type="secondary", use_container_width=True,
                                 key="stop_auto_match"):
                        st.session_state["auto_match_running"] = False
                        st.session_state.pop("auto_match_candidates", None)
                        logger.info("[自动匹配] 用户手动停止")
                        st.rerun()
                else:
                    btn_label = f"🚀 开始匹配（{len(selected_ids)}个岗位）" if selected_ids else "🚀 开始匹配"
                    if st.button(btn_label, type="primary", use_container_width=True,
                                 key="start_auto_match",
                                 disabled=not selected_ids or not ai_service.is_configured):
                        st.session_state["auto_match_running"] = True
                        st.session_state["auto_match_queue"] = selected_ids
                        st.session_state["auto_match_done"] = 0
                        st.session_state["auto_match_total"] = len(selected_ids)
                        logger.info("[自动匹配] 开始，共 {} 个岗位: {}", len(selected_ids), selected_ids)
                        st.rerun()
            with btn_cols[1]:
                if st.button("🗑️ 清除所有匹配", use_container_width=True,
                             key="clear_all_matches", disabled=is_running):
                    st.session_state["_all_positions"] = positions
                    _confirm_clear_all_matches()
            # 缩小左栏岗位 expander 内的 markdown / caption 字体（保留 streamlit 原生排版）
            st.markdown(
                """<style>
                div[data-testid="stExpander"] .stMarkdown p,
                div[data-testid="stExpander"] .stMarkdown li,
                div[data-testid="stExpander"] .stMarkdown h1,
                div[data-testid="stExpander"] .stMarkdown h2,
                div[data-testid="stExpander"] .stMarkdown h3,
                div[data-testid="stExpander"] [data-testid="stCaptionContainer"],
                div[data-testid="stExpander"] [data-testid="stCaptionContainer"] p {
                    font-size: 13px !important;
                    line-height: 1.5 !important;
                }
                </style>""",
                unsafe_allow_html=True,
            )
            for pos in positions:
                is_sel = pos.position_id == st.session_state.match_selected_pos
                exp_label = ("▶ " if is_sel else "") + pos.title
                if pos.salary_range:
                    exp_label += f"（{pos.salary_range}）"
                with st.expander(exp_label, expanded=is_sel):
                    if not is_sel:
                        # 用专属 container（带唯一 key）包按钮，再用 has-selector 锚定容器内的按钮
                        view_btn_container = st.container(key=f"viewbtn_wrap_{pos.position_id}")
                        st.markdown(
                            f"""<style>
                            div.st-key-viewbtn_wrap_{pos.position_id} button,
                            div.st-key-viewbtn_wrap_{pos.position_id} button:focus,
                            div.st-key-viewbtn_wrap_{pos.position_id} button:active {{
                                background-color: var(--color-success-soft) !important;
                                color: var(--color-success) !important;
                                border-color: var(--color-success) !important;
                            }}
                            div.st-key-viewbtn_wrap_{pos.position_id} button p {{
                                color: var(--color-success) !important;
                            }}
                            div.st-key-viewbtn_wrap_{pos.position_id} button:hover {{
                                background-color: var(--color-success) !important;
                                border-color: var(--color-success) !important;
                                color: var(--color-surface) !important;
                            }}
                            div.st-key-viewbtn_wrap_{pos.position_id} button:hover p {{
                                color: var(--color-surface) !important;
                            }}
                            </style>""",
                            unsafe_allow_html=True,
                        )
                        with view_btn_container:
                            if st.button("查看匹配结果 >>>", key=f"pos_view_{pos.position_id}",
                                         use_container_width=True):
                                st.session_state.match_selected_pos = pos.position_id
                                st.rerun()
                    meta_bits = []
                    if pos.department:
                        meta_bits.append(f"部门：{pos.department}")
                    if pos.min_education:
                        meta_bits.append(f"学历：{pos.min_education}")
                    if pos.min_experience:
                        meta_bits.append(f"年限：{pos.min_experience}")
                    if meta_bits:
                        st.caption(" | ".join(meta_bits))
                    if pos.responsibilities:
                        st.markdown(f"**岗位职责：**\n\n{pos.responsibilities}")
                    if pos.job_requirements:
                        st.markdown(f"**任职要求：**\n\n{pos.job_requirements}")
                    if not pos.responsibilities and not pos.job_requirements:
                        st.caption("（无岗位职责/任职要求说明）")

                    pos_btn_cols = st.columns(2)
                    if pos_btn_cols[0].button("🎯 智能匹配", key=f"pos_match_{pos.position_id}",
                                               type="primary", use_container_width=True,
                                               disabled=not ai_service.is_configured):
                        st.session_state.match_selected_pos = pos.position_id
                        st.session_state["trigger_match"] = pos.position_id
                        st.rerun()
                    if pos_btn_cols[1].button("🧹 清除匹配", key=f"pos_clear_{pos.position_id}",
                                               use_container_width=True):
                        st.session_state.match_selected_pos = pos.position_id
                        session_tmp = create_resume_session()
                        ResumeArchiveService(session_tmp).clear_position_matches(pos.position_id)
                        session_tmp.close()
                        st.rerun()
                    pos_btn_cols2 = st.columns(2)
                    if pos_btn_cols2[0].button("✏️ 编辑岗位", key=f"pos_edit_{pos.position_id}",
                                               use_container_width=True):
                        st.session_state["edit_pos_id"] = pos.position_id
                        _open_edit_position_dialog()
                    if pos_btn_cols2[1].button("🗑️ 删除", key=f"pos_del_{pos.position_id}",
                                               use_container_width=True):
                        job_svc.delete_position(pos.position_id)
                        st.session_state.pop("match_selected_pos", None)
                        pg_session.close()
                        st.rerun()

        # ---- 右栏：只放匹配的候选人简历 ----
        def _score_color(score: int) -> str:
            """50% → 红棕, 95%+ → 亮绿, HSL 连续渐变。"""
            clamped = max(50, min(score, 95))
            hue = 15 + (clamped - 50) * 105 / 45
            return f"hsl({hue:.0f}, 70%, 38%)"

        with match_col:
            sel_pos = next((p for p in positions if p.position_id == st.session_state.match_selected_pos), None)
            if not sel_pos:
                pg_session.close()
            else:
                # SQLite session for candidate data and position matching
                resume_session = create_resume_session()
                svc = ResumeArchiveService(resume_session)

                # ---- 自动匹配逻辑（逐岗位 rerun 模式）----
                if st.session_state.get("auto_match_running"):
                    queue = st.session_state.get("auto_match_queue", [])
                    done = st.session_state.get("auto_match_done", 0)
                    total = st.session_state.get("auto_match_total", 1)

                    st.progress(done / total if total > 0 else 0)

                    if queue:
                        current_id = queue[0]
                        current_pos = next((p for p in positions if p.position_id == current_id), None)

                        try:
                            if current_pos:
                                st.info(f"正在匹配 {done + 1}/{total}：{current_pos.title}")
                                if "auto_match_candidates" not in st.session_state:
                                    all_cands, _ = svc.list_candidates(page=1, page_size=10000)
                                    st.session_state["auto_match_candidates"] = _build_candidate_dicts(all_cands)

                                cand_dicts = st.session_state["auto_match_candidates"]
                                save_ok, save_fail, timed_out = _run_single_position_match(
                                    current_pos, svc, ai_service, cand_dicts
                                )
                                logger.info("[自动匹配] {}/{} 完成: {} — 保存 {} 条, 失败 {} 条{}",
                                            done + 1, total, current_pos.title, save_ok, save_fail,
                                            ", 超时截断" if timed_out else "")
                                if save_fail > 0:
                                    st.warning(f"岗位「{current_pos.title}」有 {save_fail} 条匹配结果保存失败。调试日志已保存到 logs/match_debug/ 目录")
                                if timed_out:
                                    st.warning(f"岗位「{current_pos.title}」匹配超时（>180s），已保存部分结果，继续下一个岗位")
                        except Exception as exc:
                            logger.error("[自动匹配] 岗位 '{}' 异常终止: {}", current_pos.title if current_pos else current_id, exc)
                            st.error(f"岗位匹配出错：{exc}")

                        st.session_state["auto_match_queue"] = queue[1:]
                        st.session_state["auto_match_done"] = done + 1
                        st.rerun()
                    else:
                        st.session_state["auto_match_running"] = False
                        st.session_state.pop("auto_match_candidates", None)
                        logger.info("[自动匹配] 全部完成，共 {} 个岗位", total)
                        st.balloons()
                        st.rerun()

                # 如果左栏触发了匹配，执行 AI 评估
                if st.session_state.pop("trigger_match", None) == sel_pos.position_id:
                    all_candidates, _ = svc.list_candidates(page=1, page_size=10000)
                    candidate_dicts = _build_candidate_dicts(all_candidates)
                    logger.info("[手动匹配] 岗位='{}' (id={}), 候选人={}, 候选人摘要={}",
                                sel_pos.title, sel_pos.position_id,
                                len(all_candidates), len(candidate_dicts))

                    # ✨ 使用统一的匹配函数（包含调试日志）
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    status_text.markdown(f"AI 正在评估候选人匹配度…")

                    save_ok, save_fail, timed_out = _run_single_position_match(
                        sel_pos, svc, ai_service, candidate_dicts
                    )

                    progress_bar.empty()
                    status_text.empty()

                    if save_fail > 0:
                        st.warning(f"AI 匹配完成：保存 {save_ok} 条，失败 {save_fail} 条"
                                   + (f"（匹配超时）" if timed_out else "")
                                   + "\n\n💡 **提示**：调试日志已保存在 logs/match_debug/ 目录，可用于远程排查问题")
                    else:
                        st.success(f"AI 匹配完成：保存 {save_ok} 条"
                                   + (f"（匹配超时）" if timed_out else ""))
                    st.rerun()

                # 展示匹配结果 Banner
                matches = svc.list_position_matches(sel_pos.position_id, min_score=30)
                if matches:
                    st.markdown(f"**{sel_pos.title}** — 匹配结果（≥30%，共 {len(matches)} 人）")
                    for match_row, cand in matches:
                        score = match_row.score
                        reason = match_row.reason or ""
                        with st.container(border=True):
                            # 第一行：姓名+主信息 左顶头 | 匹配度右侧 | 邀约按钮
                            h_cols = st.columns([3.5, 1.5, 0.3, 1.2])
                            info_parts = []
                            if cand.gender:
                                info_parts.append(cand.gender)
                            if cand.age:
                                info_parts.append(f"{cand.age}岁")
                            if cand.education_level:
                                info_parts.append(cand.education_level)
                            if cand.current_city:
                                info_parts.append(cand.current_city)
                            if cand.work_experiences and cand.work_experiences[0].position:
                                info_parts.append(cand.work_experiences[0].position)
                            info_str = f"&nbsp;&nbsp;&nbsp;{'  |  '.join(info_parts)}" if info_parts else ""
                            h_cols[0].markdown(
                                f"<div style='margin:0; line-height:1.4; padding-top:12px;'>"
                                f"<span style='font-size:22px; font-weight:700; color:var(--color-text);'>{cand.name}</span>"
                                f"<span style='font-size:14px; color:var(--color-text-secondary);'>{info_str}</span></div>",
                                unsafe_allow_html=True,
                            )
                            color = _score_color(score)
                            h_cols[1].markdown(
                                f"<div style='text-align:right;'>"
                                f"<span style='font-size:13px; color:{color}; font-weight:600;'>匹配度 {score}%</span>"
                                f"</div>",
                                unsafe_allow_html=True,
                            )

                            # ✨ 多维度评分展示
                            if hasattr(match_row, 'skill_match') and match_row.skill_match is not None:
                                dim_cols = st.columns(4)
                                dimensions = [
                                    ("技能", match_row.skill_match),
                                    ("经验", match_row.experience_match),
                                    ("学历", match_row.education_match),
                                    ("地域", match_row.location_match),
                                ]
                                for idx, (label, score_val) in enumerate(dimensions):
                                    if score_val is not None:
                                        dim_cols[idx].metric(label, f"{score_val}%", delta=None)

                            # 第 3 列空白做气口
                            with h_cols[3]:
                                st.markdown("<div style='padding-top:10px;'></div>",
                                            unsafe_allow_html=True)
                                if st.button("📧 邀约面试", key=f"match_invite_{sel_pos.position_id}_{cand.candidate_id}"):
                                    st.session_state["invite_dialog_cid"] = cand.candidate_id
                                    _open_invite_dialog()

                            # AI 评语：主题主色（紧贴上方）
                            if reason:
                                st.markdown(
                                    f"<div style='color:var(--color-primary); margin:0;'>"
                                    f"💡 <b>AI 评语：</b>{reason}</div>",
                                    unsafe_allow_html=True,
                                )

                            # 简要履历
                            edu_str = " / ".join(
                                f"{e.school_name}({e.education_level or ''}·{e.major or ''})"
                                for e in (cand.educations or [])[:2]
                            )
                            work_str = " → ".join(
                                f"{w.company_name}·{w.position or '-'}（{_fmt_date(w.start_date) or '?'} 至 {_fmt_date(w.end_date) or '至今'}）"
                                for w in (cand.work_experiences or [])[:3]
                            )
                            if edu_str:
                                st.markdown(f"🎓 {edu_str}")
                            if work_str:
                                st.markdown(f"💼 {work_str}")

                            # 联系方式（黑色正常字号）
                            contacts = []
                            if cand.phone:
                                contacts.append(f"📱 {cand.phone}")
                            if cand.email:
                                contacts.append(f"📧 {cand.email}")
                            if cand.wechat:
                                contacts.append(f"💬 {cand.wechat}")
                            contact_text = (" | ".join(contacts) if contacts else "无联系方式")
                            st.markdown(
                                f"<div style='color:var(--color-text); font-size:14px; margin:4px 0;'>{contact_text}</div>",
                                unsafe_allow_html=True,
                            )

                            # 简历文件 + 右下角来源信息
                            if cand.resume_source and cand.resume_source.file_path:
                                fp = Path(cand.resume_source.file_path)
                                file_btn_key = f"filebtns_{sel_pos.position_id}_{cand.candidate_id}"
                                st.markdown(
                                    f"""<style>
                                    .st-key-{file_btn_key} {{
                                        margin-top: -8px !important;
                                        margin-bottom: 0 !important;
                                    }}
                                    .st-key-{file_btn_key} [data-testid="stHorizontalBlock"] {{
                                        gap: 4px !important;
                                        align-items: center !important;
                                    }}
                                    .st-key-{file_btn_key} button {{
                                        min-width: 32px !important;
                                        width: 32px !important;
                                        height: 32px !important;
                                        padding: 0 !important;
                                        display: inline-flex !important;
                                        align-items: center !important;
                                        justify-content: center !important;
                                        border-radius: 6px !important;
                                        font-size: 14px !important;
                                        line-height: 1 !important;
                                    }}
                                    .st-key-{file_btn_key} button p {{
                                        margin: 0 !important;
                                        padding: 0 !important;
                                        font-size: 14px !important;
                                        line-height: 1 !important;
                                    }}
                                    </style>""",
                                    unsafe_allow_html=True,
                                )
                                src_info = ""
                                if cand.resume_source:
                                    rs = cand.resume_source
                                    src_info = f"来源：{rs.source_platform or '-'} | 入库：{rs.crawl_time.strftime('%Y-%m-%d') if rs.crawl_time else '-'}"
                                with st.container(key=file_btn_key):
                                    file_cols = st.columns([7, 0.6, 0.6, 5])
                                    file_cols[0].markdown(
                                        f"<div style='color:var(--color-text); font-size:13px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; padding-top:6px;'>📎 {fp}</div>",
                                        unsafe_allow_html=True,
                                    )
                                    if file_cols[1].button("📄", key=f"mopen_{sel_pos.position_id}_{cand.candidate_id}", help="打开文件"):
                                        try:
                                            os.startfile(str(fp))
                                        except OSError:
                                            pass
                                    if file_cols[2].button("📁", key=f"mdir_{sel_pos.position_id}_{cand.candidate_id}", help="访问目录"):
                                        try:
                                            os.startfile(str(fp.parent))
                                        except OSError:
                                            pass
                                    file_cols[3].markdown(
                                        f"<div style='text-align:right; color:var(--color-text-muted); font-size:12px; padding-top:8px;'>{src_info}</div>",
                                        unsafe_allow_html=True,
                                    )
                else:
                    st.info("暂无匹配结果。在左侧岗位中点击「🎯 智能匹配」开始 AI 评估。")

                resume_session.close()
                pg_session.close()
